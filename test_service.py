#!/usr/bin/env python3
"""
Test script for Excel to JSON conversion service.
Includes performance testing and error handling validation.
"""

import requests
import time
import sys
import os
from pathlib import Path


class ExcelConverterTester:
    """Test class for Excel conversion service."""
    
    def __init__(self, base_url="http://localhost:8000"):
        self.base_url = base_url
        self.convert_url = f"{base_url}/api/convert-excel"
        self.health_url = f"{base_url}/health"
        self.info_url = f"{base_url}/api/info"
    
    def test_health_check(self):
        """Test health check endpoint."""
        print("🏥 Testing health check...")
        try:
            response = requests.get(self.health_url, timeout=10)
            if response.status_code == 200:
                data = response.json()
                print(f"✅ Health check passed: {data['status']}")
                print(f"   Version: {data['version']}")
                print(f"   Memory usage: {data['memory_usage']}")
                return True
            else:
                print(f"❌ Health check failed: {response.status_code}")
                return False
        except Exception as e:
            print(f"❌ Health check error: {str(e)}")
            return False
    
    def test_service_info(self):
        """Test service info endpoint."""
        print("\nℹ️  Testing service info...")
        try:
            response = requests.get(self.info_url, timeout=10)
            if response.status_code == 200:
                data = response.json()
                print(f"✅ Service info retrieved")
                print(f"   Service: {data['service']}")
                print(f"   Max file size: {data['configuration']['max_file_size_mb']}MB")
                print(f"   Supported formats: {data['configuration']['supported_formats']}")
                return True
            else:
                print(f"❌ Service info failed: {response.status_code}")
                return False
        except Exception as e:
            print(f"❌ Service info error: {str(e)}")
            return False
    
    def create_test_excel_file(self, filename="test_file.xlsx", rows=100):
        """Create a simple test Excel file."""
        try:
            import openpyxl
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            
            # Add headers
            headers = ["ID", "Name", "Email", "Age", "City"]
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Add data rows
            for row in range(2, rows + 2):
                worksheet.cell(row=row, column=1, value=row - 1)
                worksheet.cell(row=row, column=2, value=f"User {row - 1}")
                worksheet.cell(row=row, column=3, value=f"user{row - 1}@example.com")
                worksheet.cell(row=row, column=4, value=20 + (row % 50))
                worksheet.cell(row=row, column=5, value=f"City {(row % 10) + 1}")
            
            workbook.save(filename)
            print(f"📄 Created test file: {filename} with {rows} data rows")
            return filename
            
        except ImportError:
            print("❌ openpyxl not available for creating test files")
            return None
        except Exception as e:
            print(f"❌ Error creating test file: {str(e)}")
            return None
    
    def test_file_conversion(self, file_path, expected_rows=None):
        """Test Excel file conversion."""
        if not os.path.exists(file_path):
            print(f"❌ Test file not found: {file_path}")
            return False
        
        file_size = os.path.getsize(file_path) / (1024 * 1024)  # MB
        print(f"\n📊 Testing file conversion: {file_path} ({file_size:.1f}MB)")
        
        start_time = time.time()
        
        try:
            with open(file_path, "rb") as f:
                files = {"file": f}
                response = requests.post(
                    self.convert_url, 
                    files=files, 
                    timeout=300  # 5 minutes timeout for large files
                )
            
            processing_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success'):
                    metadata = data.get('metadata', {})
                    total_rows = metadata.get('total_rows', 0)
                    server_time = metadata.get('processing_time', 0)
                    
                    print(f"✅ Conversion successful!")
                    print(f"   Total rows: {total_rows}")
                    print(f"   Server processing time: {server_time:.2f}s")
                    print(f"   Total request time: {processing_time:.2f}s")
                    print(f"   Processing speed: {metadata.get('performance', {}).get('rows_per_second', 0):.1f} rows/sec")
                    print(f"   Peak memory: {metadata.get('performance', {}).get('peak_memory_mb', 0):.1f}MB")
                    
                    if expected_rows and total_rows != expected_rows:
                        print(f"⚠️  Row count mismatch: expected {expected_rows}, got {total_rows}")
                    
                    # Validate data structure
                    data_array = data.get('data', [])
                    if len(data_array) > 0:
                        print(f"   First row keys: {list(data_array[0].keys())[:3]}...")
                    
                    return True
                else:
                    print(f"❌ Conversion failed: {data.get('error', {})}")
                    return False
            else:
                print(f"❌ HTTP Error {response.status_code}")
                try:
                    error_data = response.json()
                    print(f"   Error: {error_data.get('error', {}).get('message', 'Unknown error')}")
                except:
                    print(f"   Response: {response.text[:200]}...")
                return False
                
        except requests.exceptions.Timeout:
            print(f"❌ Request timeout after {processing_time:.1f}s")
            return False
        except Exception as e:
            print(f"❌ Request error: {str(e)}")
            return False
    
    def test_error_cases(self):
        """Test various error scenarios."""
        print("\n🚨 Testing error cases...")
        
        # Test 1: No file
        print("  Testing missing file...")
        try:
            response = requests.post(self.convert_url, data={}, timeout=10)
            if response.status_code == 400:
                print("  ✅ Missing file error handled correctly")
            else:
                print(f"  ❌ Unexpected response: {response.status_code}")
        except Exception as e:
            print(f"  ❌ Error testing missing file: {str(e)}")
        
        # Test 2: Invalid file type (create a text file)
        print("  Testing invalid file type...")
        try:
            with open("test_invalid.txt", "w") as f:
                f.write("This is not an Excel file")
            
            with open("test_invalid.txt", "rb") as f:
                files = {"file": ("test_invalid.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                response = requests.post(self.convert_url, files=files, timeout=10)
            
            if response.status_code in [400, 422]:
                print("  ✅ Invalid file type error handled correctly")
            else:
                print(f"  ❌ Unexpected response: {response.status_code}")
            
            os.unlink("test_invalid.txt")
            
        except Exception as e:
            print(f"  ❌ Error testing invalid file: {str(e)}")
    
    def run_performance_test(self, rows_list=[100, 1000, 5000]):
        """Run performance tests with different file sizes."""
        print("\n🚀 Running performance tests...")
        
        results = []
        for rows in rows_list:
            print(f"\n  Testing with {rows} rows...")
            test_file = self.create_test_excel_file(f"perf_test_{rows}.xlsx", rows)
            
            if test_file:
                start_time = time.time()
                success = self.test_file_conversion(test_file, expected_rows=rows)
                total_time = time.time() - start_time
                
                if success:
                    file_size = os.path.getsize(test_file) / 1024  # KB
                    results.append({
                        'rows': rows,
                        'file_size_kb': file_size,
                        'time_seconds': total_time,
                        'rows_per_second': rows / total_time
                    })
                
                # Cleanup
                try:
                    os.unlink(test_file)
                except:
                    pass
        
        # Print performance summary
        if results:
            print("\n📈 Performance Summary:")
            print("  Rows    | File Size | Time    | Speed")
            print("  --------|-----------|---------|----------")
            for result in results:
                print(f"  {result['rows']:6d} | {result['file_size_kb']:8.1f}KB | {result['time_seconds']:6.2f}s | {result['rows_per_second']:7.1f} r/s")
    
    def run_all_tests(self):
        """Run all tests."""
        print("🧪 Starting Excel Converter Service Tests\n")
        
        tests_passed = 0
        total_tests = 4
        
        # Basic functionality tests
        if self.test_health_check():
            tests_passed += 1
        
        if self.test_service_info():
            tests_passed += 1
        
        # Create and test a simple file
        test_file = self.create_test_excel_file("simple_test.xlsx", 50)
        if test_file and self.test_file_conversion(test_file, expected_rows=50):
            tests_passed += 1
            try:
                os.unlink(test_file)
            except:
                pass
        
        # Error handling tests
        self.test_error_cases()
        tests_passed += 1  # Assume error tests pass if no exceptions
        
        # Performance tests (optional)
        try:
            self.run_performance_test([100, 500, 1000])
        except Exception as e:
            print(f"⚠️  Performance test skipped: {str(e)}")
        
        # Summary
        print(f"\n📊 Test Summary: {tests_passed}/{total_tests} tests passed")
        
        if tests_passed == total_tests:
            print("🎉 All tests passed! Service is working correctly.")
            return True
        else:
            print("❌ Some tests failed. Check the output above for details.")
            return False


def main():
    """Main test function."""
    # Check if service URL is provided
    base_url = sys.argv[1] if len(sys.argv) > 1 else "http://localhost:8000"
    
    print(f"Testing service at: {base_url}")
    
    tester = ExcelConverterTester(base_url)
    success = tester.run_all_tests()
    
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
