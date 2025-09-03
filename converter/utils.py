"""
Utilities for Excel file processing with streaming and batching support.
"""
import logging
import os
import tempfile
import time
from typing import Iterator, List, Dict, Any, Tuple
import psutil
import openpyxl
import xlrd
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger('converter')


class ExcelProcessingError(Exception):
    """Custom exception for Excel processing errors."""
    pass


class MemoryMonitor:
    """Monitor memory usage during processing."""
    
    def __init__(self):
        self.process = psutil.Process()
        self.initial_memory = self.get_memory_usage()
        self.peak_memory = self.initial_memory
    
    def get_memory_usage(self) -> float:
        """Get current memory usage in MB."""
        return self.process.memory_info().rss / (1024 * 1024)
    
    def update_peak(self):
        """Update peak memory usage."""
        current = self.get_memory_usage()
        if current > self.peak_memory:
            self.peak_memory = current
    
    def get_memory_info(self) -> Dict[str, float]:
        """Get memory usage information."""
        current = self.get_memory_usage()
        return {
            'current_mb': current,
            'peak_mb': self.peak_memory,
            'initial_mb': self.initial_memory,
            'increase_mb': current - self.initial_memory
        }


def batch_rows(rows: Iterator, batch_size: int = 1000) -> Iterator[List]:
    """
    Batch rows into chunks for memory-efficient processing.
    
    Args:
        rows: Iterator of rows
        batch_size: Number of rows per batch
        
    Yields:
        List of rows in each batch
    """
    batch = []
    for row in rows:
        batch.append(row)
        if len(batch) >= batch_size:
            yield batch
            batch = []
    
    # Yield remaining rows
    if batch:
        yield batch


def process_cell_value(cell_value: Any) -> Any:
    """
    Process individual cell value to ensure proper JSON serialization.
    
    Args:
        cell_value: Raw cell value from Excel
        
    Returns:
        Processed value suitable for JSON
    """
    if cell_value is None:
        return None
    
    # Handle different cell types
    if isinstance(cell_value, (int, float)):
        # Keep numbers as numbers
        return cell_value
    elif isinstance(cell_value, str):
        # Clean up string values
        return cell_value.strip() if cell_value else None
    else:
        # Convert other types to string
        str_value = str(cell_value)
        return str_value.strip() if str_value else None


def get_column_headers(worksheet, file_type: str) -> Tuple[List[str], List[str], int]:
    """
    Extract column headers from the Excel file.
    For files with title rows, we need to find the actual headers.
    
    Args:
        worksheet: Excel worksheet object
        file_type: Type of file ('xlsx' or 'xls')
        
    Returns:
        Tuple of (title_row_headers, actual_headers, max_columns)
    """
    title_headers = []
    actual_headers = []
    max_columns = 0
    
    try:
        if file_type == 'xlsx':
            # For openpyxl (xlsx files)
            # Get first row (might be title)
            first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            
            # Get second row (likely actual headers)
            try:
                second_row = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))
            except StopIteration:
                second_row = first_row  # Only one row
            
            # Process headers with correct __EMPTY numbering
            title_headers = []
            actual_headers = []
            empty_counter = 0
            
            for i, (title_cell, actual_cell) in enumerate(zip(first_row, second_row)):
                # For title headers (used as keys)
                try:
                    if title_cell is not None and str(title_cell).strip():
                        title_headers.append(str(title_cell))
                    else:
                        if empty_counter == 0:
                            title_headers.append("__EMPTY")
                        else:
                            title_headers.append(f"__EMPTY_{empty_counter}")
                        empty_counter += 1
                except Exception:
                    # If any error in processing title cell, treat as empty
                    if empty_counter == 0:
                        title_headers.append("__EMPTY")
                    else:
                        title_headers.append(f"__EMPTY_{empty_counter}")
                    empty_counter += 1
                
                # For actual headers (used as values in first object)
                try:
                    if actual_cell is not None:
                        actual_headers.append(str(actual_cell))
                    else:
                        actual_headers.append(None)
                except Exception:
                    # If any error in processing actual cell, use None
                    actual_headers.append(None)
            
            max_columns = worksheet.max_column
            # Handle case where max_column is None
            if max_columns is None:
                max_columns = max(len(title_headers), len(actual_headers)) if (title_headers or actual_headers) else 1
        else:
            # For xlrd (xls files)
            if worksheet.nrows > 0:
                first_row = worksheet.row_values(0)
                second_row = worksheet.row_values(1) if worksheet.nrows > 1 else first_row
                
                # Process headers with correct __EMPTY numbering
                title_headers = []
                actual_headers = []
                empty_counter = 0
                
                for i, (title_cell, actual_cell) in enumerate(zip(first_row, second_row)):
                    # For title headers (used as keys)
                    try:
                        if title_cell and str(title_cell).strip():
                            title_headers.append(str(title_cell))
                        else:
                            if empty_counter == 0:
                                title_headers.append("__EMPTY")
                            else:
                                title_headers.append(f"__EMPTY_{empty_counter}")
                            empty_counter += 1
                    except Exception:
                        # If any error in processing title cell, treat as empty
                        if empty_counter == 0:
                            title_headers.append("__EMPTY")
                        else:
                            title_headers.append(f"__EMPTY_{empty_counter}")
                        empty_counter += 1
                    
                    # For actual headers (used as values in first object)
                    try:
                        if actual_cell:
                            actual_headers.append(str(actual_cell))
                        else:
                            actual_headers.append(None)
                    except Exception:
                        # If any error in processing actual cell, use None
                        actual_headers.append(None)
                    
                max_columns = worksheet.ncols
    
    except Exception as e:
        logger.error(f"Error extracting headers: {str(e)}")
        raise ExcelProcessingError(f"Failed to extract column headers: {str(e)}")
    
    # Determine which row contains the actual headers
    # If the second row looks more like headers (contains text like "NO.", "NAME", etc.)
    # and first row looks like a title, use the structure accordingly
    
    # For the format you showed, we want:
    # - First object: title_header -> actual_header mapping
    # - Subsequent objects: title_header -> data mapping
    
    # Don't trim headers - we'll handle empty columns dynamically per row
    # This allows each row to have different column structures
    
    # Ensure we have at least some headers
    if not title_headers and not actual_headers:
        raise ExcelProcessingError("No headers found in Excel file")
    
    # Handle duplicate headers by adding suffixes for title headers
    seen_headers = {}
    unique_title_headers = []
    for i, header in enumerate(title_headers):
        if header in seen_headers:
            seen_headers[header] += 1
            unique_title_headers.append(f"__EMPTY" if i == 0 else f"__EMPTY_{i}")
        else:
            seen_headers[header] = 0
            unique_title_headers.append(header)
    
    return unique_title_headers, actual_headers, max_columns


def process_xlsx_file(file_path: str, batch_size: int = 1000) -> Iterator[Dict[str, Any]]:
    """
    Process XLSX file using openpyxl with streaming support.
    
    Args:
        file_path: Path to the XLSX file
        batch_size: Number of rows to process in each batch
        
    Yields:
        Dictionary containing batch processing results
    """
    memory_monitor = MemoryMonitor()
    start_time = time.time()
    
    try:
        # Load workbook in read-only mode for memory efficiency
        logger.info(f"Loading XLSX file: {file_path}")
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        
        # Get total rows for progress tracking
        total_rows = worksheet.max_row
        logger.info(f"Total rows detected: {total_rows}")
        
        # Handle cases where max_row is None
        if total_rows is None:
            # Try to estimate by iterating through rows to find the last non-empty row
            total_rows = 0
            try:
                for row in worksheet.iter_rows():
                    if any(cell.value is not None for cell in row):
                        total_rows = row[0].row
            except:
                total_rows = 1000  # Default estimate for progress calculation
        
        logger.info(f"Effective total rows: {total_rows}")
        
        # Extract headers from first and second rows
        title_headers, actual_headers, max_columns = get_column_headers(worksheet, 'xlsx')
        logger.info(f"Headers extracted: {len(title_headers)} columns")
        
        # Yield header information first
        yield {
            'type': 'headers',
            'data': title_headers,
            'actual_headers': actual_headers,
            'total_rows': total_rows,
            'max_columns': max_columns
        }
        
        # Process data rows in batches - start from row 1 to capture all data
        row_iterator = worksheet.iter_rows(min_row=1, values_only=True)
        batch_count = 0
        total_processed = 0
        row_number = 0
        
        for batch in batch_rows(row_iterator, batch_size):
            batch_count += 1
            memory_monitor.update_peak()
            
            # Process batch
            batch_data = []
            for row_values in batch:
                row_number += 1
                if row_values and any(cell is not None for cell in row_values):
                    # Create row dictionary - only include columns with data up to the last meaningful column
                    row_dict = {}
                    
                    # Find the last column with actual data in this row
                    last_data_col = -1
                    for i, cell in enumerate(row_values):
                        if cell is not None and str(cell).strip():
                            last_data_col = i
                    
                    # Only process columns up to the last one with data
                    if last_data_col >= 0:
                        for i in range(last_data_col + 1):
                            cell_value = row_values[i] if i < len(row_values) else None
                            processed_value = process_cell_value(cell_value)
                            
                            # Only include columns that have actual data (not null/empty)
                            if processed_value is not None and str(processed_value).strip():
                                # Generate column name dynamically for each row
                                # Check if this column position has a meaningful header from the first row
                                if i < len(title_headers) and title_headers[i] and not title_headers[i].startswith('__EMPTY'):
                                    # Use the actual header from first row
                                    col_name = title_headers[i]
                                else:
                                    # Generate __EMPTY names based on actual column position
                                    if i == 0:
                                        col_name = "__EMPTY"
                                    else:
                                        col_name = f"__EMPTY_{i}"
                                
                                row_dict[col_name] = processed_value
                    
                    if row_dict:  # Only add non-empty rows
                        batch_data.append(row_dict)
                
                total_processed += 1
            
            # Calculate progress (handle case where total_rows might still be problematic)
            if total_rows and total_rows > 1:
                progress_percentage = (total_processed / max(total_rows - 1, 1)) * 100
            else:
                progress_percentage = min(100.0, (total_processed / max(batch_count * batch_size, 1)) * 100)
            processing_time = time.time() - start_time
            
            logger.info(
                f"Processed batch {batch_count}: {len(batch_data)} rows, "
                f"Progress: {progress_percentage:.1f}%, "
                f"Memory: {memory_monitor.get_memory_info()['current_mb']:.1f}MB"
            )
            
            yield {
                'type': 'batch',
                'data': batch_data,
                'batch_number': batch_count,
                'batch_size': len(batch_data),
                'total_processed': total_processed,
                'progress_percentage': progress_percentage,
                'processing_time': processing_time,
                'memory_info': memory_monitor.get_memory_info()
            }
        
        # Final summary
        total_time = time.time() - start_time
        final_memory = memory_monitor.get_memory_info()
        
        logger.info(
            f"XLSX processing completed: {total_processed} rows in {total_time:.2f}s, "
            f"Peak memory: {final_memory['peak_mb']:.1f}MB"
        )
        
        yield {
            'type': 'summary',
            'total_processed': total_processed,
            'total_batches': batch_count,
            'processing_time': total_time,
            'memory_info': final_memory,
            'rows_per_second': total_processed / total_time if total_time > 0 else 0
        }
        
    except InvalidFileException as e:
        logger.error(f"Invalid XLSX file: {str(e)}")
        raise ExcelProcessingError(f"Invalid Excel file format: {str(e)}")
    except Exception as e:
        logger.error(f"Error processing XLSX file: {str(e)}")
        raise ExcelProcessingError(f"Failed to process XLSX file: {str(e)}")
    finally:
        try:
            workbook.close()
        except:
            pass


def process_xls_file(file_path: str, batch_size: int = 1000) -> Iterator[Dict[str, Any]]:
    """
    Process XLS file using xlrd with batching support.
    
    Args:
        file_path: Path to the XLS file
        batch_size: Number of rows to process in each batch
        
    Yields:
        Dictionary containing batch processing results
    """
    memory_monitor = MemoryMonitor()
    start_time = time.time()
    
    try:
        # Open XLS file
        logger.info(f"Loading XLS file: {file_path}")
        workbook = xlrd.open_workbook(file_path)
        worksheet = workbook.sheet_by_index(0)  # Use first sheet
        
        total_rows = worksheet.nrows
        logger.info(f"Total rows detected: {total_rows}")
        
        if total_rows == 0:
            raise ExcelProcessingError("Excel file is empty")
        
        # Extract headers from first and second rows
        title_headers, actual_headers, max_columns = get_column_headers(worksheet, 'xls')
        logger.info(f"Headers extracted: {len(title_headers)} columns")
        
        # Yield header information first
        yield {
            'type': 'headers',
            'data': title_headers,
            'actual_headers': actual_headers,
            'total_rows': total_rows,
            'max_columns': max_columns
        }
        
        # Process data rows in batches (skip first 2 rows)
        batch_count = 0
        total_processed = 0
        
        for start_row in range(2, total_rows, batch_size):
            batch_count += 1
            memory_monitor.update_peak()
            
            end_row = min(start_row + batch_size, total_rows)
            batch_data = []
            
            for row_idx in range(start_row, end_row):
                try:
                    row_values = worksheet.row_values(row_idx)
                    if row_values and any(cell != '' for cell in row_values):
                        # Create row dictionary - only include columns with data up to the last meaningful column
                        row_dict = {}
                        
                        # Find the last column with actual data in this row
                        last_data_col = -1
                        for i, cell in enumerate(row_values):
                            if cell is not None and str(cell).strip():
                                last_data_col = i
                        
                        # Only process columns up to the last one with data
                        if last_data_col >= 0:
                            for i in range(last_data_col + 1):
                                cell_value = row_values[i] if i < len(row_values) else None
                                processed_value = process_cell_value(cell_value)
                                
                                # Only include columns that have actual data (not null/empty)
                                if processed_value is not None and str(processed_value).strip():
                                    # Generate column name dynamically for each row
                                    # Check if this column position has a meaningful header from the first row
                                    if i < len(title_headers) and title_headers[i] and not title_headers[i].startswith('__EMPTY'):
                                        # Use the actual header from first row
                                        col_name = title_headers[i]
                                    else:
                                        # Generate __EMPTY names based on actual column position
                                        if i == 0:
                                            col_name = "__EMPTY"
                                        else:
                                            col_name = f"__EMPTY_{i}"
                                    
                                    row_dict[col_name] = processed_value
                        
                        if row_dict:  # Only add non-empty rows
                            batch_data.append(row_dict)
                    
                    total_processed += 1
                except Exception as e:
                    logger.warning(f"Error processing row {row_idx}: {str(e)}")
                    continue
            
            # Calculate progress
            progress_percentage = (total_processed / max(total_rows - 1, 1)) * 100
            processing_time = time.time() - start_time
            
            logger.info(
                f"Processed batch {batch_count}: {len(batch_data)} rows, "
                f"Progress: {progress_percentage:.1f}%, "
                f"Memory: {memory_monitor.get_memory_info()['current_mb']:.1f}MB"
            )
            
            yield {
                'type': 'batch',
                'data': batch_data,
                'batch_number': batch_count,
                'batch_size': len(batch_data),
                'total_processed': total_processed,
                'progress_percentage': progress_percentage,
                'processing_time': processing_time,
                'memory_info': memory_monitor.get_memory_info()
            }
        
        # Final summary
        total_time = time.time() - start_time
        final_memory = memory_monitor.get_memory_info()
        
        logger.info(
            f"XLS processing completed: {total_processed} rows in {total_time:.2f}s, "
            f"Peak memory: {final_memory['peak_mb']:.1f}MB"
        )
        
        yield {
            'type': 'summary',
            'total_processed': total_processed,
            'total_batches': batch_count,
            'processing_time': total_time,
            'memory_info': final_memory,
            'rows_per_second': total_processed / total_time if total_time > 0 else 0
        }
        
    except Exception as e:
        logger.error(f"Error processing XLS file: {str(e)}")
        raise ExcelProcessingError(f"Failed to process XLS file: {str(e)}")


def process_excel_streaming(uploaded_file, batch_size: int = 1000) -> Dict[str, Any]:
    """
    Main function to process Excel file with streaming support.
    
    Args:
        uploaded_file: Django UploadedFile object
        batch_size: Number of rows to process in each batch
        
    Returns:
        Dictionary containing the complete processed data and metadata
    """
    start_time = time.time()
    temp_file_path = None
    
    try:
        # Determine file type from extension
        file_name = uploaded_file.name.lower()
        if file_name.endswith('.xlsx'):
            file_type = 'xlsx'
        elif file_name.endswith('.xls'):
            file_type = 'xls'
        else:
            raise ExcelProcessingError("Unsupported file format. Only .xlsx and .xls files are supported.")
        
        # Save uploaded file to temporary location
        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_type}') as temp_file:
            temp_file_path = temp_file.name
            for chunk in uploaded_file.chunks():
                temp_file.write(chunk)
        
        logger.info(
            f"Processing {file_type.upper()} file: {uploaded_file.name}, "
            f"Size: {uploaded_file.size / (1024 * 1024):.1f}MB"
        )
        
        # Process file based on type
        if file_type == 'xlsx':
            processor = process_xlsx_file(temp_file_path, batch_size)
        else:
            processor = process_xls_file(temp_file_path, batch_size)
        
        # Collect all data
        all_data = []
        title_headers = []
        actual_headers = []
        summary_info = {}
        
        for result in processor:
            if result['type'] == 'headers':
                title_headers = result['data']
                actual_headers = result['actual_headers']
                
                # For this type of file, we don't add a header mapping row
                # The first row of actual data will be the first object
                # This handles files where each row has different content structure
                
            elif result['type'] == 'batch':
                all_data.extend(result['data'])
            elif result['type'] == 'summary':
                summary_info = result
        
        # Calculate final metadata
        total_time = time.time() - start_time
        file_size_mb = uploaded_file.size / (1024 * 1024)
        
        metadata = {
            'total_rows': len(all_data) - 1,  # Subtract header row
            'processing_time': round(total_time, 3),
            'file_size': f"{file_size_mb:.1f}MB",
            'worksheet_name': 'Sheet1',  # Default name, could be enhanced
            'file_type': file_type.upper(),
            'batch_size': batch_size,
            'performance': {
                'rows_per_second': round(summary_info.get('rows_per_second', 0), 2),
                'peak_memory_mb': round(summary_info.get('memory_info', {}).get('peak_mb', 0), 1),
                'total_batches': summary_info.get('total_batches', 0)
            }
        }
        
        logger.info(
            f"Excel processing completed successfully: "
            f"{metadata['total_rows']} rows, {total_time:.2f}s, "
            f"{metadata['performance']['rows_per_second']} rows/sec"
        )
        
        return {
            'success': True,
            'data': all_data,
            'metadata': metadata
        }
        
    except ExcelProcessingError:
        raise
    except Exception as e:
        logger.error(f"Unexpected error during Excel processing: {str(e)}")
        raise ExcelProcessingError(f"Unexpected error: {str(e)}")
    finally:
        # Clean up temporary file
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except Exception as e:
                logger.warning(f"Failed to delete temporary file {temp_file_path}: {str(e)}")
